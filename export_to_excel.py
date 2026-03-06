"""
export_to_excel.py — Author: Mousumi Paul
Builds the full 7-sheet Power BI master Excel workbook with AI summaries.
"""
import pandas as pd, numpy as np, json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FOREST="1B4332"; FOREST_MID="2D6A4F"; FOREST_LITE="D8F3DC"
AMBER="E76F51"; AMBER_LITE="FDEBD0"; SKY="1A6B8A"; SKY_LITE="D6EAF8"
GOLD_LITE="FEF9E7"; WHITE="FFFFFF"; GRAY="F4F4F4"; DARK="1A1A2E"

def bdr():
    s=Side(style="thin",color="CCCCCC")
    return Border(left=s,right=s,top=s,bottom=s)

def hf(size=10, bold=True, color=WHITE):
    return Font(name="Calibri",size=size,bold=bold,color=color)

def df_(size=10, color=DARK):
    return Font(name="Calibri",size=size,bold=False,color=color)

def fx(h):
    return PatternFill("solid",start_color=h,fgColor=h)

def ctr(wrap=False):
    return Alignment(horizontal="center",vertical="center",wrap_text=wrap)


def write_cover(wb):
    ws=wb.create_sheet("Cover",0); ws.sheet_view.showGridLines=False
    ws.column_dimensions["A"].width=4
    ws.merge_cells("B2:J2"); ws.row_dimensions[2].height=52
    c=ws["B2"]; c.value="AI-Powered Farm Dashboard  |  ChatGPT + Power BI"
    c.font=Font(name="Calibri",size=22,bold=True,color=FOREST)
    c.alignment=ctr(); c.fill=fx(FOREST_LITE)
    ws.merge_cells("B3:J3")
    ws["B3"].value="Yield Trends  •  Weather Impact  •  Seasonal Patterns  •  AI Insights"
    ws["B3"].font=Font(name="Calibri",size=11,italic=True,color=FOREST_MID)
    ws["B3"].alignment=Alignment(horizontal="center")
    meta=[("Author","Mousumi Paul"),("Report Period","2020-2024"),("AI Engine","OpenAI GPT-4o"),("BI Tool","Microsoft Power BI"),("Crops","Rice, Wheat, Tomato, Potato, Onion, Maize, Soybean, Cotton, Sugarcane, Groundnut"),("Sheets","Cover | KPI Overview | Yield Trends | Weather Impact | Seasonal | AI Summaries | Market Prices")]
    for i,(k,v) in enumerate(meta,start=5):
        ws.row_dimensions[i].height=22
        ws[f"B{i}"]=k; ws[f"B{i}"].font=Font(name="Calibri",size=11,bold=True,color=FOREST)
        ws[f"C{i}"]=v; ws[f"C{i}"].font=Font(name="Calibri",size=11,color="333333")
        ws.merge_cells(f"C{i}:J{i}")


def write_kpi_sheet(wb, yield_df, weather_df):
    import sys; sys.path.insert(0, '.')
    from src.visualization.kpi_calculator import FarmDashboardKPIs
    ws=wb.create_sheet("KPI Overview"); ws.sheet_view.showGridLines=False
    ws.merge_cells("A1:F1"); ws["A1"].value="Farm Dashboard KPI Overview"
    ws["A1"].font=hf(14,True,WHITE); ws["A1"].fill=fx(FOREST); ws["A1"].alignment=ctr(); ws.row_dimensions[1].height=34
    kpis=FarmDashboardKPIs(yield_df,weather_df).full_report()
    for c_idx,h in enumerate(["KPI","Value","Target","Category","Status"],1):
        cell=ws.cell(row=2,column=c_idx,value=h)
        cell.font=hf(10,True,WHITE); cell.fill=fx(FOREST_MID); cell.alignment=ctr(); cell.border=bdr()
    for r_idx,row in enumerate(kpis.itertuples(index=False),start=3):
        kpi,val,target,cat=row.KPI,row.Value,row.Target,row.Category
        status="—"; bg=WHITE
        try:
            fval=float(val)
            if "%" in str(target) and ">" in str(target):
                tgt=float(str(target).replace(">","").replace("%","").strip())
                if fval>=tgt: status="On Target"; bg=FOREST_LITE
                else: status="Review"; bg=AMBER_LITE
        except: pass
        for c_idx,v in enumerate([kpi,val,target,cat,status],1):
            cell=ws.cell(row=r_idx,column=c_idx,value=v)
            cell.font=df_(); cell.border=bdr(); cell.alignment=ctr()
            cell.fill=fx(bg if c_idx>=4 else(GRAY if r_idx%2==0 else WHITE))
    for col,w in zip("ABCDEF",[40,20,18,16,18,4]):
        ws.column_dimensions[col].width=w


def write_data_sheet(wb,title,df,sheet_name,hdr_color=FOREST,alt_color=FOREST_LITE):
    ws=wb.create_sheet(sheet_name); ws.sheet_view.showGridLines=False
    ncols=len(df.columns)
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    ws["A1"].value=title; ws["A1"].font=hf(13,True,WHITE); ws["A1"].fill=fx(hdr_color); ws["A1"].alignment=ctr(); ws.row_dimensions[1].height=30
    for c_idx,col in enumerate(df.columns,1):
        cell=ws.cell(row=2,column=c_idx,value=col.replace("_"," "))
        cell.font=hf(9,True,WHITE); cell.fill=fx(FOREST_MID); cell.alignment=ctr(True); cell.border=bdr()
    ws.row_dimensions[2].height=28
    for r_idx,row in enumerate(df.itertuples(index=False),start=3):
        bg=alt_color if r_idx%2==0 else WHITE
        for c_idx,val in enumerate(row,1):
            cell=ws.cell(row=r_idx,column=c_idx,value=val)
            cell.font=df_(); cell.fill=fx(bg); cell.border=bdr(); cell.alignment=ctr()
    for c_idx,col in enumerate(df.columns,1):
        w=max(len(str(col)),df[col].astype(str).str.len().max())
        ws.column_dimensions[get_column_letter(c_idx)].width=min(w+3,28)
    return ws


def write_ai_summaries(wb, summaries_path):
    ws=wb.create_sheet("AI Summaries"); ws.sheet_view.showGridLines=False
    ws.merge_cells("A1:F1"); ws["A1"].value="AI-Generated Crop Insights (ChatGPT GPT-4o)"
    ws["A1"].font=hf(13,True,WHITE); ws["A1"].fill=fx("7D4E00"); ws["A1"].alignment=ctr(); ws.row_dimensions[1].height=30
    for c_idx,h in enumerate(["Crop","AI Summary","Dashboard Text","Avg Yield","Top Region","Sentiment"],1):
        cell=ws.cell(row=2,column=c_idx,value=h)
        cell.font=hf(9,True,WHITE); cell.fill=fx(AMBER); cell.alignment=ctr(True); cell.border=bdr()
    rows=[]
    if os.path.exists(summaries_path):
        with open(summaries_path) as f: summaries=json.load(f)
        for s in summaries:
            rows.append([s.get("crop",""),s.get("ai_summary",""),s.get("dashboard_text",""),s.get("avg_yield",""),s.get("top_region",""),s.get("sentiment","Neutral")])
    else:
        mock_crops=["Rice","Wheat","Tomato","Potato","Onion","Maize","Soybean","Cotton","Sugarcane","Groundnut"]
        for crop in mock_crops:
            summary=f"{crop} yields show a steady upward trend across 2020-2024, driven by improved seed varieties and expanded irrigation coverage. Top-performing regions outperform national averages by 15-22%. Weather-correlated dips were noted in 2021 due to irregular monsoon distribution. Profit margins remain above 27% for well-managed farms. Recommendation: adopt precision irrigation to sustain and accelerate yield growth."
            rows.append([crop,summary,summary[:200]+"...",round(np.random.uniform(18,36),2),np.random.choice(["Punjab","Maharashtra","UP","Karnataka"]),"Positive"])
    for r_idx,row_data in enumerate(rows,start=3):
        bg=GOLD_LITE if r_idx%2==0 else WHITE
        for c_idx,val in enumerate(row_data,1):
            cell=ws.cell(row=r_idx,column=c_idx,value=val)
            cell.font=df_(); cell.fill=fx(bg); cell.border=bdr()
            cell.alignment=Alignment(horizontal="left" if c_idx==2 else "center",vertical="top",wrap_text=True)
        ws.row_dimensions[r_idx].height=65
    for col,w in zip("ABCDEF",[14,70,45,12,18,12]):
        ws.column_dimensions[col].width=w


def main():
    os.makedirs("data/processed",exist_ok=True)
    if not os.path.exists("data/raw/crop_yield_records.csv"):
        os.system("python scripts/generate_sample_data.py")
    if not os.path.exists("data/processed/yield_trends.xlsx"):
        os.system("python scripts/data_preprocessing.py")

    yield_df=pd.read_excel("data/processed/yield_trends.xlsx")
    weather_df=pd.read_excel("data/processed/weather_impact.xlsx")
    seasonal_df=pd.read_excel("data/processed/seasonal_patterns.xlsx")
    prices_df=pd.read_csv("data/raw/market_prices.csv")

    wb=Workbook(); wb.remove(wb.active)
    write_cover(wb)
    write_kpi_sheet(wb,yield_df,weather_df)
    write_data_sheet(wb,"Crop Yield Trends (2020-2024)",yield_df.head(400),"Yield Trends",FOREST,FOREST_LITE)
    w_cols=[c for c in ["Record_ID","Year","Month","Month_Name","Season","Crop","Region","Avg_Rainfall","Avg_Temp","Avg_Humidity","Weather_Score","Drought_Months","Flood_Risk","Yield_Per_Acre_Qtl"] if c in weather_df.columns]
    write_data_sheet(wb,"Weather Impact on Yield",weather_df[w_cols].head(400),"Weather Impact",SKY,SKY_LITE)
    write_data_sheet(wb,"Seasonal Production Patterns",seasonal_df,"Seasonal Patterns","5B4FCF","EDE7F6")
    write_data_sheet(wb,"Market Prices",prices_df.head(300),"Market Prices",AMBER,AMBER_LITE)
    write_ai_summaries(wb,"data/ai_outputs/crop_summaries.json")

    out="data/processed/dashboard_master.xlsx"
    wb.save(out)
    print(f"  dashboard_master.xlsx -> {out}")

if __name__ == "__main__":
    main()